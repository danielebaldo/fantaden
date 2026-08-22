"""Esporta la board in export/lista_asta.xlsx: la classica lista Excel di
scelte e fasce, un foglio per ruolo, ordinata per fascia e poi per score,
pronta da stampare o consultare durante l'asta.

Uso:
    python3 scripts/export_xlsx.py
"""
import json
import os

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lib.config import load_settings, repo_path

ROLE_ORDER = ["P", "D", "C", "A"]
ROLE_SHEET_NAMES = {"P": "Portieri", "D": "Difensori", "C": "Centrocampisti", "A": "Attaccanti"}

FASCIA_FILL = {
    "Top": "FFD966",
    "1a Fascia": "A9D18E",
    "2a Fascia": "9DC3E6",
    "3a Fascia": "D9D9D9",
    "Low Cost": "F4CCCC",
    "Scommessa": "EAD1DC",
}

BASE_COLUMNS = [
    ("name", "Nome"),
    ("position_mantra", "Ruolo M"),
    ("team", "Squadra"),
    ("qt_att", "Qt.A"),
    ("fvm", "FVM"),
    ("fvm_500", "FVM 500cr"),
    ("fascia", "Fascia"),
    ("score", "Score"),
    ("affare_label", "Affare"),
    ("presenze", "Pv"),
    ("media_voto", "Mv"),
    ("fantamedia", "Fm"),
    ("stagioni_disponibili", "Stag."),
    ("trend_fantamedia", "Trend Fm"),
    ("continuita", "Continuità"),
    ("gol_fatti", "Gf"),
    ("assist", "Ass"),
    ("rigori_calciati", "Rc"),
    ("presenze_corrente", "Pv (corrente)"),
    ("fantamedia_corrente", "Fm (corrente)"),
    ("gol_fatti_corrente", "Gf (corrente)"),
    ("assist_corrente", "Ass (corrente)"),
]
GK_COLUMNS = [("gol_subiti", "Gs"), ("rigori_parati", "Rp")]
OUTFIELD_COLUMNS = [("ammonizioni", "Amm"), ("espulsioni", "Esp")]
EXTRA_COLUMNS = [("stelle", "Stelle"), ("note", "Note")]


def _columns_for_role(role: str):
    cols = list(BASE_COLUMNS)
    if role == "P":
        cols += GK_COLUMNS
    else:
        cols += OUTFIELD_COLUMNS
    cols += EXTRA_COLUMNS
    return cols


def _sheet_dataframe(board: list, role: str) -> pd.DataFrame:
    cols = _columns_for_role(role)
    rows = [p for p in board if p["position"] == role]
    data = []
    for p in rows:
        row = {label: p.get(key) for key, label in cols}
        data.append(row)
    df = pd.DataFrame(data, columns=[label for _, label in cols])
    df["Appunti"] = ""  # colonna vuota per note scritte a mano durante l'asta
    return df


def _style_sheet(ws, df: pd.DataFrame, board_rows: list):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="404040")
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    fascia_col_idx = list(df.columns).index("Fascia") + 1 if "Fascia" in df.columns else None
    if fascia_col_idx:
        for row_idx, p in enumerate(board_rows, start=2):
            color = FASCIA_FILL.get(p.get("fascia"))
            if color:
                ws.cell(row=row_idx, column=fascia_col_idx).fill = PatternFill("solid", fgColor=color)

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].tolist()])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 8), 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main():
    settings = load_settings()
    board_path = repo_path(settings["paths"]["board_json"])
    export_path = repo_path(settings["paths"]["export_xlsx"])

    if not os.path.exists(board_path):
        raise SystemExit(f"[export_xlsx] Manca {board_path}: esegui prima scripts/build_board.py")

    with open(board_path, "r", encoding="utf-8") as f:
        board = json.load(f)

    os.makedirs(os.path.dirname(export_path), exist_ok=True)
    with pd.ExcelWriter(export_path, engine="openpyxl") as writer:
        for role in ROLE_ORDER:
            df = _sheet_dataframe(board, role)
            sheet_name = ROLE_SHEET_NAMES[role]
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            role_rows = [p for p in board if p["position"] == role]
            _style_sheet(ws, df, role_rows)

    print(f"[export_xlsx] scritto {export_path} ({len(board)} giocatori, {len(ROLE_ORDER)} fogli)")


if __name__ == "__main__":
    main()
